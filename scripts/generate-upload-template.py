#!/usr/bin/env python3
"""
Generate Template CSV untuk Upload Massal JelajahBelanja
Format: .xlsx (bisa dibuka di Excel, save as CSV untuk upload)
"""

import sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─── Color System ───
PRIMARY = "1B2A4A"
PRIMARY_LIGHT = "D6E4F0"
NEUTRAL_900 = "37352F"
NEUTRAL_600 = "8C8A84"
NEUTRAL_200 = "E9E9E8"
NEUTRAL_100 = "F7F7F5"
NEUTRAL_0 = "FFFFFF"
ACCENT_POSITIVE = "1B7D46"
ACCENT_WARNING = "D4820A"

# ─── Fonts ───
FONT_NAME = "Calibri"
font_title = Font(name=FONT_NAME, size=16, bold=True, color=PRIMARY)
font_header = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
font_body = Font(name=FONT_NAME, size=11, color=NEUTRAL_900)
font_caption = Font(name=FONT_NAME, size=9, color=NEUTRAL_600)
font_required_header = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
font_optional_header = Font(name=FONT_NAME, size=11, bold=False, color=PRIMARY)
font_example = Font(name=FONT_NAME, size=11, color=NEUTRAL_600, italic=True)
font_note = Font(name=FONT_NAME, size=10, color=ACCENT_WARNING, bold=True)

# ─── Fills ───
fill_header = PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type="solid")
fill_header_optional = PatternFill(start_color=PRIMARY_LIGHT, end_color=PRIMARY_LIGHT, fill_type="solid")
fill_alt = PatternFill(start_color=NEUTRAL_100, end_color=NEUTRAL_100, fill_type="solid")
fill_example = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")

# ─── Borders ───
border_header_bottom = Border(bottom=Side(style="thin", color=NEUTRAL_200))

# ─── Alignments ───
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center")

wb = Workbook()

# ═══════════════════════════════════════════════════
# SHEET 1: Template Upload (data input)
# ═══════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Template Upload"
ws1.sheet_properties.tabColor = PRIMARY

# Column A = margin
ws1.column_dimensions["A"].width = 3

# ─── Title ───
ws1.merge_cells("B2:O2")
ws1["B2"].value = "Template Upload Massal — JelajahBelanja"
ws1["B2"].font = font_title
ws1["B2"].alignment = Alignment(horizontal="left", vertical="center")
ws1.row_dimensions[2].height = 32

# ─── Subtitle ───
ws1.merge_cells("B3:O3")
ws1["B3"].value = "Isi data produk di bawah, lalu Save As → CSV (Comma Separated Values) untuk upload massal"
ws1["B3"].font = font_caption
ws1["B3"].alignment = Alignment(horizontal="left", vertical="center")
ws1.row_dimensions[3].height = 18

# ─── Headers ───
# Kolom wajib dan opsional
columns = [
    # (header, width, required, description, example)
    ("title", 35, True, "Nama produk", "Celana Jeans Pria Slim Fit"),
    ("url", 40, True, "Link produk di marketplace", "https://shopee.co.id/product-123"),
    ("image", 40, True, "Link gambar produk", "https://example.com/image1.jpg"),
    ("price", 16, True, "Harga jual (angka, tanpa Rp)", 150000),
    ("category", 16, True, "Kategori produk", "Fashion"),
    ("originalPrice", 18, False, "Harga asli sebelum diskon", 250000),
    ("discountPercent", 18, False, "Persentase diskon", 40),
    ("rating", 12, False, "Rating 0-5 (kosongkan kalau belum ada review)", 4.8),
    ("reviewCount", 16, False, "Jumlah review (0 kalau belum ada)", 1200),
    ("soldCount", 16, False, "Jumlah terjual", 5000),
    ("location", 18, False, "Lokasi seller", "Jakarta"),
    ("marketplace", 16, False, "shopee / tokopedia / lazada / dll", "shopee"),
    ("affiliateUrl", 35, False, "Link affiliate (kalau ada)", "https://shope.ee/aff123"),
    ("notes", 25, False, "Catatan internal", ""),
]

header_row = 5
ws1.row_dimensions[4].height = 8  # spacer

for i, (header, width, required, desc, example) in enumerate(columns):
    col = i + 2  # start from B
    letter = get_column_letter(col)

    # Column width
    ws1.column_dimensions[letter].width = width

    # Header cell
    cell = ws1.cell(row=header_row, column=col, value=header)
    if required:
        cell.font = font_required_header
        cell.fill = fill_header
    else:
        cell.font = font_optional_header
        cell.fill = fill_header_optional
    cell.alignment = align_center
    cell.border = border_header_bottom

# Header row height
ws1.row_dimensions[header_row].height = 28

# ─── Description row (penjelasan tiap kolom) ───
desc_row = header_row + 1
ws1.row_dimensions[desc_row].height = 30
for i, (header, width, required, desc, example) in enumerate(columns):
    col = i + 2
    cell = ws1.cell(row=desc_row, column=col, value=desc)
    cell.font = Font(name=FONT_NAME, size=9, color=ACCENT_WARNING if required else NEUTRAL_600, italic=True)
    cell.alignment = align_center
    cell.fill = PatternFill(start_color="FFFDE7" if required else NEUTRAL_100, end_color="FFFDE7" if required else NEUTRAL_100, fill_type="solid")

# ─── Example rows (3 contoh produk) ───
example_data = [
    ["Celana Jeans Pria Slim Fit", "https://shopee.co.id/product-123", "https://example.com/image1.jpg", 150000, "Fashion", 250000, 40, 4.8, 1200, 5000, "Jakarta", "shopee", "https://shope.ee/aff123", ""],
    ["Kaos Oversize Unisex", "https://shopee.co.id/product-456", "https://example.com/image2.jpg", 75000, "Fashion", 120000, 38, 4.7, 800, 3000, "Bandung", "shopee", "", ""],
    ["Sepatu Sneakers Pria", "https://shopee.co.id/product-789", "https://example.com/image3.jpg", 250000, "Fashion", "", "", 0, 0, 0, "Surabaya", "shopee", "", "produk baru belum ada rating"],
]

for row_idx, row_data in enumerate(example_data):
    row = desc_row + 1 + row_idx
    ws1.row_dimensions[row].height = 22
    for col_idx, value in enumerate(row_data):
        cell = ws1.cell(row=row, column=col_idx + 2, value=value)
        cell.font = font_example
        cell.fill = fill_example
        cell.alignment = align_right if isinstance(value, (int, float)) else align_left

# ─── Empty rows for input (20 baris) ───
input_start = desc_row + 1 + len(example_data)
for row_idx in range(20):
    row = input_start + row_idx
    ws1.row_dimensions[row].height = 22
    fill = fill_alt if row_idx % 2 == 1 else PatternFill(fill_type=None)
    for col_idx in range(len(columns)):
        cell = ws1.cell(row=row, column=col_idx + 2)
        cell.font = font_body
        cell.fill = fill
        cell.alignment = align_left

# ─── Data Validation: marketplace ───
mp_col = next(i for i, c in enumerate(columns) if c[0] == "marketplace") + 2
mp_letter = get_column_letter(mp_col)
dv_marketplace = DataValidation(
    type="list",
    formula1='"shopee,tokopedia,lazada,aliexpress,amazon"',
    allow_blank=True,
    showDropDown=False
)
dv_marketplace.prompt = "Pilih marketplace"
dv_marketplace.promptTitle = "Marketplace"
ws1.add_data_validation(dv_marketplace)
dv_marketplace.add(f"{mp_letter}{input_start}:{mp_letter}{input_start + 19}")

# ─── Data Validation: rating 0-5 ───
rating_col = next(i for i, c in enumerate(columns) if c[0] == "rating") + 2
rating_letter = get_column_letter(rating_col)
dv_rating = DataValidation(
    type="decimal",
    operator="between",
    formula1=0,
    formula2=5,
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Rating tidak valid",
    error="Rating harus antara 0 - 5"
)
ws1.add_data_validation(dv_rating)
dv_rating.add(f"{rating_letter}{input_start}:{rating_letter}{input_start + 19}")

# ─── Freeze panes ───
ws1.freeze_panes = f"B{input_start}"

# ─── Legend row ───
legend_row = input_start + 21
ws1.row_dimensions[legend_row].height = 8
legend_row2 = legend_row + 1
ws1.merge_cells(f"B{legend_row2}:O{legend_row2}")
ws1[f"B{legend_row2}"].value = "LEGENDA:"
ws1[f"B{legend_row2}"].font = Font(name=FONT_NAME, size=10, bold=True, color=PRIMARY)

legend_row3 = legend_row2 + 1
ws1.merge_cells(f"B{legend_row3}:O{legend_row3}")
ws1[f"B{legend_row3}"].value = "Kolom header gelap = WAJIB diisi  |  Kolom header terang = OPSIONAL  |  Baris kuning = contoh data (hapus sebelum upload)"
ws1[f"B{legend_row3}"].font = font_caption

legend_row4 = legend_row3 + 1
ws1.merge_cells(f"B{legend_row4}:O{legend_row4}")
ws1[f"B{legend_row4}"].value = "Untuk produk yang belum punya rating/review, biarkan rating=0 dan reviewCount=0 (sistem akan tampilkan 'Belum ada rating')"
ws1[f"B{legend_row4}"].font = font_caption

legend_row5 = legend_row4 + 1
ws1.merge_cells(f"B{legend_row5}:O{legend_row5}")
ws1[f"B{legend_row5}"].value = "Hapus baris contoh (kuning) sebelum upload! Maksimal 200 produk per upload."
ws1[f"B{legend_row5}"].font = Font(name=FONT_NAME, size=10, color=ACCENT_WARNING, bold=True)

# ═══════════════════════════════════════════════════
# SHEET 2: Panduan
# ═══════════════════════════════════════════════════
ws2 = wb.create_sheet("Panduan")
ws2.sheet_properties.tabColor = ACCENT_POSITIVE

ws2.column_dimensions["A"].width = 3
ws2.column_dimensions["B"].width = 5
ws2.column_dimensions["C"].width = 80

# Title
ws2.merge_cells("B2:C2")
ws2["B2"].value = "Panduan Upload Massal JelajahBelanja"
ws2["B2"].font = font_title
ws2.row_dimensions[2].height = 32

# Steps
steps = [
    ("1.", "Buka tab 'Template Upload', isi data produk di baris kosong (setelah baris contoh kuning)"),
    ("2.", "Kolom WAJIB (header gelap): title, url, image, price, category — harus diisi semua"),
    ("3.", "Kolom OPSIONAL (header terang): boleh dikosongkan, sistem akan pakai default"),
    ("4.", "Untuk produk BARU yang belum punya review: isi rating=0, reviewCount=0"),
    ("5.", "Hapus baris contoh (warna kuning) sebelum upload!"),
    ("6.", "Save As → pilih format 'CSV (Comma Separated Values) (*.csv)'"),
    ("7.", "Buka admin JelajahBelanja → tab 'Upload Massal' → drag & drop file CSV"),
    ("8.", "Cek preview, lalu klik Upload — semua produk masuk sekali jalan!"),
    ("9.", "Maksimal 200 produk per upload"),
]

for i, (num, text) in enumerate(steps):
    row = 4 + i
    ws2.row_dimensions[row].height = 24
    ws2[f"B{row}"].value = num
    ws2[f"B{row}"].font = Font(name=FONT_NAME, size=11, bold=True, color=PRIMARY)
    ws2[f"B{row}"].alignment = align_center
    ws2[f"C{row}"].value = text
    ws2[f"C{row}"].font = font_body
    ws2[f"C{row}"].alignment = align_left

# Tips section
tip_start = 4 + len(steps) + 1
ws2.merge_cells(f"B{tip_start}:C{tip_start}")
ws2[f"B{tip_start}"].value = "Tips & Catatan"
ws2[f"B{tip_start}"].font = Font(name=FONT_NAME, size=13, bold=True, color=PRIMARY)
ws2.row_dimensions[tip_start].height = 28

tips = [
    "Ambil data dari halaman produk Shopee: copy link produk (url), link gambar, harga, dll",
    "Kalau produk belum punya review, JANGAN isi rating palsu — biarkan rating=0, reviewCount=0",
    "Marketplace default = shopee (kalau dikosongkan otomatis shopee)",
    "Category bebas, tapi sebaiknya pakai kategori yang sudah ada di website",
    "AffiliateUrl bisa diisi kalau kamu punya link affiliate khusus untuk produk itu",
    "Simpan file ini sebagai MASTER — nanti tinggal copy baris baru untuk produk berikutnya",
    "Bisa kerja bertahap: isi 5-10 produk dulu, save, lanjut besok, baru upload kalau udah banyak",
]

for i, tip in enumerate(tips):
    row = tip_start + 1 + i
    ws2.row_dimensions[row].height = 22
    ws2[f"B{row}"].value = "•"
    ws2[f"B{row}"].font = Font(name=FONT_NAME, size=11, color=ACCENT_POSITIVE)
    ws2[f"B{row}"].alignment = align_center
    ws2[f"C{row}"].value = tip
    ws2[f"C{row}"].font = font_body
    ws2[f"C{row}"].alignment = align_left

# Kolom referensi
ref_start = tip_start + 1 + len(tips) + 1
ws2.merge_cells(f"B{ref_start}:C{ref_start}")
ws2[f"B{ref_start}"].value = "Referensi Kolom"
ws2[f"B{ref_start}"].font = Font(name=FONT_NAME, size=13, bold=True, color=PRIMARY)
ws2.row_dimensions[ref_start].height = 28

ref_data = [
    ("title", "Wajib", "Nama produk, contoh: Celana Jeans Pria Slim Fit"),
    ("url", "Wajib", "Link produk di marketplace, contoh: https://shopee.co.id/..."),
    ("image", "Wajib", "Link gambar produk (URL), contoh: https://..."),
    ("price", "Wajib", "Harga jual dalam angka (tanpa Rp), contoh: 150000"),
    ("category", "Wajib", "Kategori, contoh: Fashion, Elektronik, dll"),
    ("originalPrice", "Opsional", "Harga asli sebelum diskon, contoh: 250000"),
    ("discountPercent", "Opsional", "Persentase diskon, contoh: 40"),
    ("rating", "Opsional", "Rating 0-5, contoh: 4.8 (isi 0 kalau belum ada review)"),
    ("reviewCount", "Opsional", "Jumlah review, contoh: 1200 (isi 0 kalau belum ada)"),
    ("soldCount", "Opsional", "Jumlah terjual, contoh: 5000"),
    ("location", "Opsional", "Lokasi seller, contoh: Jakarta"),
    ("marketplace", "Opsional", "shopee / tokopedia / lazada / aliexpress / amazon"),
    ("affiliateUrl", "Opsional", "Link affiliate khusus produk ini"),
    ("notes", "Opsional", "Catatan internal (tidak ditampilkan ke user)"),
]

for i, (col_name, req, desc) in enumerate(ref_data):
    row = ref_start + 1 + i
    ws2.row_dimensions[row].height = 20
    fill = fill_alt if i % 2 == 1 else PatternFill(fill_type=None)
    
    ws2[f"B{row}"].value = col_name
    ws2[f"B{row}"].font = Font(name=FONT_NAME, size=10, bold=True, color=PRIMARY if req == "Wajib" else NEUTRAL_600)
    ws2[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws2[f"B{row}"].fill = fill
    
    ws2[f"C{row}"].value = f"[{req}] {desc}"
    ws2[f"C{row}"].font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)
    ws2[f"C{row}"].alignment = align_left
    ws2[f"C{row}"].fill = fill


# ─── Save ───
output_path = "/home/z/my-project/download/template-upload-jelajahbelanja.xlsx"
wb.save(output_path)
print(f"✅ Template saved: {output_path}")
